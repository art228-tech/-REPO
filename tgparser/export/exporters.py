"""Выгрузка базы в CSV, XLSX и JSON."""

from __future__ import annotations

import csv
import io
import json
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from tgparser.db.models import Lead

COLUMNS: list[tuple[str, str]] = [
    ("tg_user_id", "ID"),
    ("tag", "Тег"),
    ("display_name", "Имя"),
    ("chat_title", "Чат"),
    ("topic_title", "Топик"),
    ("message_link", "Ссылка на сообщение"),
    ("archive_link", "Карточка в архиве"),
    ("archive_anonymized", "Ссылка на автора скрыта"),
    ("message_date", "Дата сообщения"),
    ("snippet", "Текст"),
    ("source", "Источник"),
    ("status", "Статус"),
    ("is_premium", "Premium"),
    ("note", "Заметка"),
    ("created_at", "Добавлен"),
]

SOURCE_LABELS = {
    "roster": "список участников",
    "history": "автор сообщения",
    "comment": "комментарий",
    "manual": "вручную",
}


@dataclass(slots=True)
class ExportResult:
    path: Path
    rows: int
    fmt: str


def _cell(lead: Lead, key: str) -> Any:
    if key == "tag":
        return lead.tag or ""
    if key == "display_name":
        return lead.display_name
    if key == "source":
        return SOURCE_LABELS.get(lead.source, lead.source)
    value = getattr(lead, key, None)
    if isinstance(value, bool):
        return "да" if value else ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return "" if value is None else value


def _rows(leads: Sequence[Lead]) -> list[list[Any]]:
    return [[_cell(lead, key) for key, _ in COLUMNS] for lead in leads]


def to_csv(
    leads: Sequence[Lead],
    path: Path,
    delimiter: str = ";",
    bom: bool = True,
) -> ExportResult:
    """CSV для Excel.

    Без BOM Excel открывает UTF-8 как ANSI и кириллица превращается в кракозябры,
    поэтому по умолчанию он есть. Разделитель `;` — Excel в русской локали
    ожидает именно его.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow([title for _, title in COLUMNS])
    writer.writerows(_rows(leads))

    encoding = "utf-8-sig" if bom else "utf-8"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(buffer.getvalue(), encoding=encoding, newline="")
    return ExportResult(path=path, rows=len(leads), fmt="csv")


def to_xlsx(leads: Sequence[Lead], path: Path) -> ExportResult:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лиды"

    titles = [title for _, title in COLUMNS]
    sheet.append(titles)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for row in _rows(leads):
        sheet.append(row)

    widths = [10, 20, 26, 28, 18, 34, 34, 14, 18, 60, 18, 12, 10, 24, 18]
    for index, width in enumerate(widths[: len(COLUMNS)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return ExportResult(path=path, rows=len(leads), fmt="xlsx")


def to_json(leads: Sequence[Lead], path: Path) -> ExportResult:
    payload = []
    for lead in leads:
        item: dict[str, Any] = {}
        for key, _ in COLUMNS:
            value = _cell(lead, key)
            if key in {"archive_anonymized", "is_premium"}:
                item[key] = bool(getattr(lead, key, False))
            else:
                item[key] = value if value != "" else None
        payload.append(item)

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return ExportResult(path=path, rows=len(leads), fmt="json")


def to_tags(leads: Sequence[Lead], path: Path) -> ExportResult:
    """Плоский список тегов — когда нужны только @ы, по одному на строку."""
    tags = [lead.tag for lead in leads if lead.tag]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(tags) + ("\n" if tags else ""), encoding="utf-8")
    return ExportResult(path=path, rows=len(tags), fmt="txt")
