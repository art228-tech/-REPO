from __future__ import annotations

import csv
import json
from datetime import UTC, datetime

import pytest

from tgparser.db.models import Lead, SourceKind
from tgparser.db.repo import LeadRepo
from tgparser.db.settings_store import ScanSettings
from tgparser.export.exporters import COLUMNS, to_csv, to_json, to_tags, to_xlsx
from tgparser.export.service import ExportFilter, build_filename, export, fetch_leads


def lead(
    lead_id: int = 1,
    username: str | None = "ivanov",
    name: str = "Пётр",
    archived: bool = False,
) -> Lead:
    item = Lead(
        id=lead_id,
        tg_user_id=1000 + lead_id,
        username=username,
        first_name=name,
        last_name="Петров",
        chat_title="Чат про станки",
        topic_title=None,
        message_link="https://t.me/c/1/50",
        message_date=datetime(2026, 7, 1, 12, 30, tzinfo=UTC),
        snippet="ищу подрядчика",
        source=SourceKind.HISTORY.value,
        status="new",
        is_premium=True,
        archive_link="https://t.me/c/9/3" if archived else None,
        archive_anonymized=archived,
        created_at=datetime(2026, 7, 2, tzinfo=UTC),
    )
    return item


class TestCsv:
    def test_writes_bom_for_excel(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead()], path, bom=True)
        assert path.read_bytes().startswith(b"\xef\xbb\xbf")

    def test_no_bom_when_disabled(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead()], path, bom=False)
        assert not path.read_bytes().startswith(b"\xef\xbb\xbf")

    def test_semicolon_delimiter(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead()], path, delimiter=";")
        header = path.read_text(encoding="utf-8-sig").splitlines()[0]
        assert header.count(";") == len(COLUMNS) - 1

    def test_cyrillic_reads_back(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead(name="Пётр")], path)
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=";"))
        assert "Пётр Петров" in rows[1]

    def test_header_matches_columns(self, tmp_path):
        path = tmp_path / "out.csv"
        result = to_csv([], path)
        assert result.rows == 0
        with path.open(encoding="utf-8-sig", newline="") as handle:
            header = next(csv.reader(handle, delimiter=";"))
        assert header == [title for _, title in COLUMNS]

    def test_tag_column_has_at_sign(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead(username="ivanov")], path)
        assert "@ivanov" in path.read_text(encoding="utf-8-sig")

    def test_untagged_leaves_empty_cell(self, tmp_path):
        path = tmp_path / "out.csv"
        to_csv([lead(username=None)], path)
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=";"))
        assert rows[1][1] == ""


class TestXlsx:
    def test_creates_readable_workbook(self, tmp_path):
        from openpyxl import load_workbook

        path = tmp_path / "out.xlsx"
        to_xlsx([lead(), lead(2, username=None)], path)

        sheet = load_workbook(path).active
        assert sheet.title == "Лиды"
        assert sheet.max_row == 3
        assert sheet.cell(row=1, column=1).value == "ID"
        assert sheet.freeze_panes == "A2"

    def test_boolean_rendered_as_word(self, tmp_path):
        from openpyxl import load_workbook

        path = tmp_path / "out.xlsx"
        to_xlsx([lead(archived=True)], path)
        sheet = load_workbook(path).active
        header = [c.value for c in sheet[1]]
        column = header.index("Ссылка на автора скрыта") + 1
        assert sheet.cell(row=2, column=column).value == "да"


class TestJson:
    def test_structure_and_encoding(self, tmp_path):
        path = tmp_path / "out.json"
        to_json([lead()], path)
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert len(payload) == 1
        assert payload[0]["tag"] == "@ivanov"
        assert payload[0]["display_name"] == "Пётр Петров"
        assert payload[0]["is_premium"] is True

    def test_empty_values_become_null(self, tmp_path):
        path = tmp_path / "out.json"
        to_json([lead(username=None)], path)
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert payload[0]["tag"] is None


class TestTags:
    def test_one_per_line(self, tmp_path):
        path = tmp_path / "out.txt"
        result = to_tags([lead(1, "one_tag"), lead(2, None), lead(3, "two_tag")], path)
        assert result.rows == 2
        assert path.read_text(encoding="utf-8").splitlines() == ["@one_tag", "@two_tag"]

    def test_empty_input(self, tmp_path):
        path = tmp_path / "out.txt"
        assert to_tags([], path).rows == 0
        assert path.read_text(encoding="utf-8") == ""


class TestFilters:
    async def test_only_with_username(self, db):
        from tgparser.db.repo import CollectedUser

        async with db.session() as session:
            repo = LeadRepo(session)
            await repo.add(CollectedUser(tg_user_id=1, username="tagged"))
            await repo.add(CollectedUser(tg_user_id=2, username=None))

        async with db.session() as session:
            found = await fetch_leads(session, ExportFilter(only_with_username=True))
        assert [item.username for item in found] == ["tagged"]

    async def test_only_without_username(self, db):
        from tgparser.db.repo import CollectedUser

        async with db.session() as session:
            repo = LeadRepo(session)
            await repo.add(CollectedUser(tg_user_id=1, username="tagged"))
            await repo.add(CollectedUser(tg_user_id=2, username=None))

        async with db.session() as session:
            found = await fetch_leads(session, ExportFilter(only_without_username=True))
        assert [item.tg_user_id for item in found] == [2]

    async def test_by_chat(self, db):
        from tgparser.db.repo import CollectedUser

        async with db.session() as session:
            repo = LeadRepo(session)
            await repo.add(CollectedUser(tg_user_id=1, username="a_tag", chat_id=-100))
            await repo.add(CollectedUser(tg_user_id=2, username="b_tag", chat_id=-200))

        async with db.session() as session:
            found = await fetch_leads(session, ExportFilter(chat_id=-200))
        assert [item.tg_user_id for item in found] == [2]

    def test_describe(self):
        assert ExportFilter().describe() == "без фильтров"
        assert "только с тегом" in ExportFilter(only_with_username=True).describe()


class TestExportService:
    async def test_rejects_unknown_format(self, db, tmp_path):
        async with db.session() as session:
            with pytest.raises(ValueError, match="Неизвестный формат"):
                await export(session, "pdf", tmp_path, ScanSettings())

    @pytest.mark.parametrize("fmt", ["csv", "xlsx", "json", "txt"])
    async def test_all_formats_produce_file(self, db, tmp_path, fmt):
        from tgparser.db.repo import CollectedUser

        async with db.session() as session:
            await LeadRepo(session).add(CollectedUser(tg_user_id=1, username="ivanov"))

        async with db.session() as session:
            result = await export(session, fmt, tmp_path, ScanSettings())

        assert result.path.exists()
        assert result.path.stat().st_size > 0

    def test_filename_has_timestamp_and_extension(self):
        name = build_filename("csv", now=datetime(2026, 8, 1, 10, 20, 30, tzinfo=UTC))
        assert name == "leads-20260801-102030.csv"
